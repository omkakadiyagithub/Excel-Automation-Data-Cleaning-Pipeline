import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
print("Creating Excel report...")

# Load datasets
clean_df = pd.read_csv("data/clean_retail_data.csv")
kpi_df = pd.read_csv("reports/kpi_summary.csv")

quality_df = pd.read_csv(
    "reports/data_quality_report.csv"
)

# Create workbook
wb = Workbook()

# Sheet 1 - Clean Data
ws1 = wb.active
ws1.title = "Clean Data"

# Headers
for col_num, column_name in enumerate(clean_df.columns, start=1):
    cell = ws1.cell(row=1, column=col_num)
    cell.value = column_name
    cell.font = Font(bold=True)

    cell.fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )
# Data
for row_num, row in enumerate(clean_df.values, start=2):
    for col_num, value in enumerate(row, start=1):
        ws1.cell(row=row_num, column=col_num, value=value)

# Sheet 2 - KPI Summary
ws2 = wb.create_sheet(title="KPI Summary")

for col_num, column_name in enumerate(kpi_df.columns, start=1):
    cell = ws2.cell(row=1, column=col_num)
    cell.value = column_name
    cell.font = Font(bold=True)

    cell.fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

for row_num, row in enumerate(kpi_df.values, start=2):
    for col_num, value in enumerate(row, start=1):
        ws2.cell(row=row_num, column=col_num, value=value)

# Data Quality Sheet

ws3 = wb.create_sheet(
    title="Data Quality Report"
)

for row in dataframe_to_rows(
    quality_df,
    index=False,
    header=True
):
    ws3.append(row)

for cell in ws3[1]:

    cell.font = Font(bold=True)

    cell.fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

# Auto-adjust column widths

for sheet in wb.worksheets:

    for column in sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2

        sheet.column_dimensions[column_letter].width = adjusted_width

# Save report
wb.save("reports/Final_Report.xlsx")

print("Final_Report.xlsx created successfully!")